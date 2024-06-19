import torch
from sklearn.metrics import confusion_matrix, classification_report, roc_auc_score, roc_curve, auc
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook

# Load probabilities and targets
probs = torch.load('probs_32.pth')
targets = torch.load('test-32.lib')['targets']

# Calculate maximum probabilities for each slide
maxs = [max(probs[i * 4624:(i + 1) * 4624]) for i in range(1046)]

# Generate predictions based on the threshold of 0.5
preds = [1 if i > 0.5 else 0 for i in maxs]

# Print confusion matrix and classification report
print(confusion_matrix(targets, preds))
print(classification_report(targets, preds, digits=4))
print('AUC:', roc_auc_score(targets, maxs))

# Calculate ROC curve
fpr, tpr, thresholds = roc_curve(targets, maxs)
roc_auc = auc(fpr, tpr)

# Save FPR and TPR to an Excel file
wb = Workbook()
ws = wb.active
for i in range(len(fpr)):
    ws.cell(row=i + 1, column=1).value = fpr[i]
    ws.cell(row=i + 1, column=2).value = tpr[i]
wb.save('result.xlsx')

# Plot and save the ROC curve
plt.figure()
lw = 2
plt.plot(fpr, tpr, color='darkorange', lw=lw, label='ROC curve (area = %0.2f)' % roc_auc)
plt.plot([0, 1], [0, 1], color='navy', lw=lw, linestyle='--')
plt.xlim([0.0, 1.0])
plt.ylim([0.0, 1.05])
plt.xlabel('False Positive Rate')
plt.ylabel('True Positive Rate')
plt.title('Receiver Operating Characteristic')
plt.legend(loc="lower right")
plt.savefig('roc_50.png')